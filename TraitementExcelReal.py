from math import *
import numpy as np
import pandas as pd
from scipy import stats
import scipy.optimize
import sys
from openpyxl import load_workbook
from FileReader import *
from sklearn.metrics import mean_squared_error, mean_absolute_error, max_error

########## Fichier de traitement pour les données réelles

## Méthodes de calcul

# Méthode pseudo-inversion directe de G : lambda = pseudoinv(G) * Pf1
# Cette fonction est équivalente à la closed-form d'Essr
def pinv_only(G_t, Pf1) :
    G = np.array(G_t).transpose()
    return np.linalg.pinv(G) @ Pf1

# Fonction de vraisemblance classique
# def likelihood(lam,reads,G):
#     lam = lam/np.sum(lam)
#     Pflambda = np.dot(G,lam)
#     vraisemblance = 1
#     for i in range(len(reads[1])):
#         proba = stats.binom.pmf(reads[1][i],reads[0][i],Pflambda[i])
#         vraisemblance = vraisemblance * proba
#     return -vraisemblance

# Fonction de vraisemblance logarithmique
def likelihoodLog(lam,reads,G):
    lam = lam/np.sum(lam)
    Pflambda = np.dot(G,lam)
    vraisemblance = 0
    for i in range(len(reads[1])):
        proba = stats.binom.pmf(reads[1][i],reads[0][i],Pflambda[i])
        if proba == 0:
            vraisemblance += sys.float_info.epsilon # Plus petit flottant possible en Python
        else :
            vraisemblance += np.log(proba)
    return -vraisemblance

## Traitement des données initiales :

def traitementDonneesReelles(fileExcel,listeMelanges):
    for melange in listeMelanges: # Nom des feuilles du excel = N° du mélange

        print("mélange N° :", melange)
        data = pd.read_excel(fileExcel,"Tm10"+str(melange).zfill(2))

        # Récupération de G
        G = pd.DataFrame(data.iloc[0:len(data.index)-2,1:len(data.columns)-1]).to_numpy()
        G_T = G.T
        nb_geno = np.shape(G_T)[1]
        nb_snps = np.shape(G_T)[0]

        # Récupération des reads
        reads = pd.DataFrame(data.iloc[[len(data.index)-2,len(data.index)-1],1:len(data.columns)-1]).to_numpy()

        # Pf1 observé qui servira à construire lam_init
        Pf1_observe = np.zeros(nb_snps)
        for i in range(nb_snps): # Obligé de tout parcourir car certaines positions ont 0 reads
            if reads[0][i] != 0:
                Pf1_observe[i] = reads[1][i] / reads[0][i]

        # Définition du lambda initial qui sera utilisé pour les optimisations
        lam_init = pinv_only(G,Pf1_observe)
        for i in range(lam_init.size):
            if lam_init[i] < 0 :
                lam_init[i]=0

        # Création des bounds pour l'utilisation dans les algorithmes
        bounds = ()
        for i in range(nb_geno):
            bounds = bounds + ((0,1),)

        ## Maximisation de la fonction de vraisemblance (minimisation de son opposé) :
        print("début calcul algorithmes")
        min_NMLog = scipy.optimize.minimize(likelihoodLog,lam_init,args=(reads,G_T),method= 'Nelder-Mead',bounds=bounds,options={'maxiter':50,'maxfev':50})
        cons = ({'type': 'eq', 'fun': lambda x: np.sum(x) - 1})
        min_SLSQPLog = scipy.optimize.minimize(likelihoodLog, lam_init, args=(reads, G_T), method='SLSQP',bounds=bounds, constraints=cons)
        print("algos ok")

        ## Traitement des résultats :

        # On crée une matrice à afficher sur excel
        lambda_differents = np.array([lam_init, min_NMLog.x, min_SLSQPLog.x])
        likelihoodsLog = np.array([likelihoodLog(lam_init,reads,G_T),likelihoodLog(min_NMLog.x,reads,G_T),likelihoodLog(min_SLSQPLog.x,reads,G_T)])
        lambda_differents = lambda_differents.T
        matrice = np.vstack((lambda_differents,likelihoodsLog))
        colonnes = ["LambdaMatrix","NMLog","SLSQPLog"]

        # Export vers Excel
        book = load_workbook(fileExcel)
        writer = pd.ExcelWriter(fileExcel,engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df = pd.DataFrame(matrice,columns=colonnes)
        df.to_excel(writer,"Tm10"+str(melange).zfill(2), startcol=3, startrow=nb_geno + 5, index=False)
        writer.save()

def traitementDonneesReelles96(fileExcel,listeFeuilles, mixtures_dict, column_nucleotypes_dict):
    print("départ")
    for feuilleNum in listeFeuilles:
        data = pd.read_excel(fileExcel,"Tm10"+str(feuilleNum).zfill(2))

        genotypes = mixtures_dict["Tm10"+str(feuilleNum).zfill(2)]
        # Récupération des fréquences initiales
        mix_indexes = []
        for genotype in genotypes :
            g_index = column_nucleotypes_dict[genotype]
            mix_indexes.append(g_index)

        print("mix indexes ", mix_indexes)

        freq = np.zeros((96, 1))
        np.put(freq, mix_indexes, np.full((12, 1), 1/12))
        print("new freq", freq)

        # Récupération de G
        G = pd.DataFrame(data.iloc[0:len(data.index) - 2,1:len(data.columns)-1]).to_numpy()
        G_T = G.T

        # Récupération des reads
        reads = pd.DataFrame(data.iloc[[len(data.index) - 2, len(data.index) - 1], 1:len(data.columns) - 1]).to_numpy()
        nb_geno = np.shape(G_T)[1]
        nb_snps = np.shape(G_T)[0]

        # Pf1 observé qui servira à construire lam_init
        Pf1_observe = np.zeros(nb_snps)
        for i in range(nb_snps):
            if reads[0][i] != 0:
                Pf1_observe[i] = reads[1][i] / reads[0][i]

        # Définition du lambda initial qui sera utilisé pour les optimisations
        lam_init = pinv_only(G, Pf1_observe)
        for i in range(lam_init.size):
            if lam_init[i] < 0:
                lam_init[i] = 0

        bounds = ()
        for i in range(nb_geno):
            bounds = bounds + ((0,1),)

        ### Maximisation de la fonction de vraisemblance (minimisation de son opposé) : tentative avec toutes les méthodes
        print("lancement algo")
        min_NMLog = scipy.optimize.minimize(likelihoodLog,lam_init,args=(reads,G_T),method= 'Nelder-Mead',bounds=bounds)
        cons = ({'type': 'eq', 'fun': lambda x: np.sum(x) - 1})
        print("NMLog ok")
        print("min lambda NMLog", min_NMLog.x)
        min_SLSQPLog = scipy.optimize.minimize(likelihoodLog,lam_init,args=(reads,G_T),method= 'SLSQP',bounds=bounds,constraints=cons)
        print("algos ok")

        # On calcule les erreurs de chaque algorithme par rapport au lambda recherché
        lambda_differents = np.array([lam_init,min_NMLog.x,min_SLSQPLog.x])
        print("lambda differents", lambda_differents)
        rmse = np.zeros(len(lambda_differents))
        mse = np.zeros(len(lambda_differents))
        mae = np.zeros(len(lambda_differents))
        maxError = np.zeros(len(lambda_differents))
        likelihoodsLog = np.zeros(len(lambda_differents))
        erreur_abs = np.zeros(len(lambda_differents))
        likeli = likelihoodLog(freq,reads,G_T)
        for i in range(len(lambda_differents)):
            lambda_algo = lambda_differents[i]
            likelihoodsLog[i] = likelihoodLog(lambda_algo,reads,G_T)
            erreurAbs = abs(freq - lambda_algo)
            erreur_abs[i] = np.mean(erreurAbs)
            rmse[i] =  sqrt(mean_squared_error(freq, lambda_algo))
            mse[i] = mean_squared_error(freq, lambda_algo)
            mae[i] = mean_absolute_error(freq, lambda_algo)
            maxError[i] = max_error(freq, lambda_algo)
        erreur_abs_min = min(erreur_abs)
        indiceMinAbs = erreur_abs.tolist().index(erreur_abs_min)
        memes_algo_concat = ["LambdaMatrix", "NMLog","SLSQPLog"]
        algoMinAbs = memes_algo_concat[indiceMinAbs]

        # On crée une matrice à afficher sur excel
        lambda_differents = lambda_differents.T
        matrice = np.vstack((lambda_differents, likelihoodsLog))
        matrice = np.vstack((matrice, erreur_abs))
        matrice = np.vstack((matrice,rmse))
        matrice = np.vstack((matrice,mse))
        matrice = np.vstack((matrice,mae))
        matrice = np.vstack((matrice,maxError))
        matrix = [erreur_abs_min, algoMinAbs,likeli]

        # On va écrire sur excel
        book = load_workbook(fileExcel)
        writer = pd.ExcelWriter(fileExcel, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df = pd.DataFrame(matrice, columns=memes_algo_concat)
        df.to_excel(writer, "N°" + str(feuilleNum), startcol=5, startrow=nb_geno + 5, index=False)
        df2 = pd.DataFrame(matrix,index = ['valeur','algorithme','likelihood_init'],columns=['erreur absolue min'])
        df2.to_excel(writer,"N°"+str(feuilleNum), startrow=nb_geno+6,startcol=1)
        writer.save()

## Filepaths
# filepath_positions = "D:/Rémi/Documents/IMT/3A/S10/EtudeTech/melange_simul_renom/positions_correspondance.txt"
# filepath_reads = "D:/Rémi/Documents/IMT/3A/S10/EtudeTech/melange_simul_renom/SIMULS_MIXTURES_fauxBAM_vraiREADS/reads_statistics.txt"
# filepath_nucleotypes = "D:/Rémi/Documents/IMT/3A/S10/EtudeTech/melange_simul_renom/nucleotypes.txt"
# filepath_mixtures = "D:/Rémi/Documents/IMT/3A/S10/EtudeTech/melange_simul_renom/simulated_mixtures_composition.txt"

filepath_positions = "C:/Users/mabed/Documents/Travail/Etudes_techniques/fichiers_travail/positions_correspondance.txt"
filepath_reads = "C:/Users/mabed/Documents/Travail/Etudes_techniques/fichiers_travail/reads_statistics.txt"
filepath_nucleotypes = "C:/Users/mabed/Documents/Travail/Etudes_techniques/fichiers_travail/nucleotypes.txt"
filepath_mixtures = "C:/Users/mabed/Documents/Travail/Etudes_techniques/fichiers_travail/simulated_mixtures_composition.txt"

## Data
data_utils = build_data_utils(filepath_positions, filepath_reads, filepath_nucleotypes, filepath_mixtures)
mixtures_dict = data_utils[0]
column_nucleotypes_dict = data_utils[2]

traitementDonneesReelles96("outputVrai.xlsx",[5,6,7], mixtures_dict, column_nucleotypes_dict)
